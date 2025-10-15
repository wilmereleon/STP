#!/usr/bin/env python3
"""
Script para crear Excel mejorado con estructura completa
"""
import sys
import os

def create_improved_excel():
    """Crea un Excel mejorado con ejemplos"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("❌ ERROR: openpyxl no instalado")
        print("📦 Instalando openpyxl...")
        os.system("pip install openpyxl")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scripts de Teleprompter"
    
    # Headers
    headers = [
        "Número de Guion",
        "Título",
        "Texto/Contenido",
        "Duración (HH:MM:SS)",
        "Notas"
    ]
    
    # Estilo para headers
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escribir headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Datos de ejemplo
    example_data = [
        {
            "numero": "001",
            "titulo": "Apertura del Noticiero",
            "texto": """Buenas noches, les saluda María González.\n\nBienvenidos a Noticias 8 PM.\n\nEstas son las principales noticias de hoy:\n\n- El gobierno anuncia nuevas medidas económicas\n- Se registra récord de vacunación en el país\n- Mañana habrá lluvia en la región norte\n\nComenzamos con la información.""",
            "duracion": "00:00:45",
            "notas": "Usar tono entusiasta. Cámara 1."
        },
        {
            "numero": "002",
            "titulo": "Noticia Economía",
            "texto": """El Ministerio de Economía anunció hoy un paquete de medidas para incentivar el empleo.\n\nEntre las principales acciones se encuentran:\n\n**Reducción de impuestos** para pequeñas empresas.\n\n*Subsidios directos* para sectores afectados por la pandemia.\n\nProgramas de capacitación gratuita.\n\nEl ministro declaró que estas medidas beneficiarán a más de 500 mil personas.\n\n[PAUSA]\n\nAhora pasamos con el reporte desde el terreno.""",
            "duracion": "00:01:30",
            "notas": "Insertar gráfico de estadísticas en segundo 20. Voz seria."
        },
        {
            "numero": "003",
            "titulo": "Reportaje Salud",
            "texto": """**REPORTAJE: CAMPAÑA DE VACUNACIÓN**\n\n*[Locutor en vivo]*\n\nHoy se alcanzó un hito histórico en la campaña de vacunación.\n\nMás de 10 millones de personas han recibido su primera dosis.\n\nLos centros de vacunación reportan alta afluencia.\n\n[CORTE A VIDEO]\n\nEntrevista con la directora de salud pública:\n\n\"Estamos muy satisfechos con la respuesta de la ciudadanía.\nLa meta es llegar al 80% de la población antes de fin de año.\"\n\n[REGRESO A ESTUDIO]\n\nUna excelente noticia para todos.""",
            "duracion": "00:02:15",
            "notas": "Video externo 45 seg. Transición suave."
        },
        {
            "numero": "004",
            "titulo": "Pronóstico del Tiempo",
            "texto": """Y ahora el clima con Roberto Martínez.\n\n*[Cambio a set del clima]*\n\nBuenas noches.\n\nEl pronóstico para mañana indica:\n\n🌤️ Norte: lluvias moderadas por la mañana\n☀️ Centro: despejado con temperatura de 25°C\n⛈️ Sur: tormentas eléctricas por la tarde\n\nRecomendamos llevar paraguas si salen temprano.\n\nLa temperatura mínima será de 18°C y la máxima de 28°C.\n\nRegresamos al estudio.""",
            "duracion": "00:01:00",
            "notas": "Animaciones del mapa. Roberto usa chroma key."
        },
        {
            "numero": "005",
            "titulo": "Cierre y Despedida",
            "texto": """Estas fueron las principales noticias de este martes.\n\nRecuerden seguirnos en nuestras redes sociales:\n\n📱 Twitter: @Noticias8PM\n📘 Facebook: Noticias 8 PM Oficial\n📷 Instagram: @noticias8pm\n\n[PAUSA LARGA]\n\nGracias por su sintonía.\n\nLes esperamos mañana a la misma hora con más información.\n\n**Buenas noches.**\n\n[MÚSICA DE CIERRE]""",
            "duracion": "00:00:40",
            "notas": "Fade out progresivo. Créditos finales."
        }
    ]
    
    # Escribir datos de ejemplo
    for row_idx, data in enumerate(example_data, 2):
        ws.cell(row_idx, 1).value = data["numero"]
        ws.cell(row_idx, 2).value = data["titulo"]
        ws.cell(row_idx, 3).value = data["texto"]
        ws.cell(row_idx, 4).value = data["duracion"]
        ws.cell(row_idx, 5).value = data["notas"]
        
        # Estilo para celdas de datos
        for col_idx in range(1, 6):
            cell = ws.cell(row_idx, col_idx)
            cell.alignment = Alignment(
                vertical='top',
                wrap_text=True
            )
            cell.border = thin_border
    
    # Agregar instrucciones en hoja separada
    ws_instrucciones = wb.create_sheet("Instrucciones")
    
    instrucciones = [
        ["INSTRUCCIONES DE USO", ""],
        ["", ""],
        ["¿Cómo usar esta plantilla?", ""],
        ["", ""],
        ["1. Número de Guion", "Identificador único del script (ej: 001, 002, 003)"],
        ["2. Título", "Nombre descriptivo del segmento"],
        ["3. Texto/Contenido", "Script completo del teleprompter. Puede incluir:"],
        ["", "  - Texto normal para el locutor"],
        ["", "  - **Negrita** para énfasis"],
        ["", "  - *Cursiva* para acotaciones"],
        ["", "  - [MARCADORES] para indicaciones técnicas"],
        ["4. Duración", "Tiempo estimado en formato HH:MM:SS"],
        ["5. Notas", "Anotaciones técnicas o recordatorios"],
        ["", ""],
        ["Formatos especiales:", ""],
        ["", ""],
        ["**Texto en negrita**", "Para palabras importantes"],
        ["*Texto en cursiva*", "Para acotaciones del director"],
        ["[MARCADOR]", "Para saltos rápidos con macros"],
        ["[PAUSA]", "Indicación de pausa dramática"],
        ["", ""],
        ["Emojis soportados:", ""],
        ["🌤️ ☀️ ⛈️ 🌡️", "Para pronósticos del tiempo"],
        ["📱 📘 📷 💻", "Para redes sociales"],
        ["🎬 🎤 🎥 📹", "Para producción"],
        ["", ""],
        ["Consejos:", ""],
        ["", ""],
        ["✅ Usar saltos de línea para facilitar lectura", ""],
        ["✅ Incluir pausas con [PAUSA] o líneas vacías", ""],
        ["✅ Marcar cambios de cámara con [CAM 1], [CAM 2]", ""],
        ["✅ Indicar transiciones con [CORTE], [FADE]", ""],
        ["✅ Resaltar nombres propios con **negrita**", ""],
        ["", ""],
        ["Importar:", ""],
        ["", ""],
        ["En el teleprompter:", ""],
        ["1. Ir a Panel de Run Order", ""],
        ["2. Click en botón 'Importar Excel'", ""],
        ["3. Seleccionar este archivo", ""],
        ["4. Vista previa y confirmar", ""],
        ["5. Todos los scripts se cargarán automáticamente", ""],
    ]
    
    for row_idx, (titulo, descripcion) in enumerate(instrucciones, 1):
        cell_a = ws_instrucciones.cell(row_idx, 1)
        cell_b = ws_instrucciones.cell(row_idx, 2)
        
        cell_a.value = titulo
        cell_b.value = descripcion
        
        # Estilo para títulos
        if titulo and not titulo.startswith(" ") and titulo != "":
            if row_idx == 1:
                cell_a.font = Font(size=16, bold=True, color='366092')
            elif titulo.endswith(":"):
                cell_a.font = Font(size=12, bold=True, color='366092')
            else:
                cell_a.font = Font(size=11, bold=True)
        
        cell_a.alignment = Alignment(vertical='top', wrap_text=True)
        cell_b.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Ajustar anchos de columna - Hoja principal
    ws.column_dimensions['A'].width = 15  # Número
    ws.column_dimensions['B'].width = 30  # Título
    ws.column_dimensions['C'].width = 80  # Texto
    ws.column_dimensions['D'].width = 18  # Duración
    ws.column_dimensions['E'].width = 35  # Notas
    
    # Ajustar anchos - Hoja instrucciones
    ws_instrucciones.column_dimensions['A'].width = 35
    ws_instrucciones.column_dimensions['B'].width = 60
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    # Guardar
    output_path = "Excel/prompter_intro_mejorado.xlsx"
    wb.save(output_path)
    
    print(f"✅ Excel mejorado creado: {output_path}")
    print(f"📊 {len(example_data)} scripts de ejemplo")
    print(f"📄 2 hojas: Scripts + Instrucciones")
    print(f"📏 Columnas configuradas con anchos óptimos")
    print(f"🎨 Formato profesional con colores y estilos")
    
    return True

if __name__ == "__main__":
    success = create_improved_excel()
    sys.exit(0 if success else 1)
