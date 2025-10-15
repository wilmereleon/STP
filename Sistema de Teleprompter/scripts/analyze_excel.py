#!/usr/bin/env python3
"""
Script para analizar y mejorar el Excel de importación
"""
import sys
import os

# Agregar parent directory al path
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

def analyze_excel():
    """Analiza el Excel existente"""
    try:
        import openpyxl
    except ImportError:
        print("❌ ERROR: openpyxl no instalado")
        print("Instalar con: pip install openpyxl")
        return False
    
    excel_path = "Excel/prompter_intro.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"❌ No se encontró: {excel_path}")
        return False
    
    wb = openpyxl.load_workbook(excel_path)
    print(f"✅ Excel cargado: {excel_path}")
    print(f"📊 Hojas: {wb.sheetnames}")
    
    ws = wb.active
    print(f"📏 Dimensiones: {ws.dimensions}")
    print(f"📋 Filas: {ws.max_row}, Columnas: {ws.max_column}")
    
    # Leer headers
    headers = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(1, col)
        headers.append(cell.value)
    
    print(f"\n📋 Headers encontrados:")
    for i, h in enumerate(headers, 1):
        print(f"  {i}. {h}")
    
    # Leer primeras 3 filas de datos
    print(f"\n📄 Primeras 3 filas de datos:")
    for row_idx in range(2, min(5, ws.max_row + 1)):
        print(f"\n  Fila {row_idx}:")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row_idx, col_idx)
            value = str(cell.value)[:50] if cell.value else "(vacío)"
            print(f"    {header}: {value}")
    
    return True

if __name__ == "__main__":
    success = analyze_excel()
    sys.exit(0 if success else 1)
